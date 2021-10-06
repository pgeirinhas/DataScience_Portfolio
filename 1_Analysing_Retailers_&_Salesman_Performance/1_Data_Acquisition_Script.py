# -*- coding: utf-8 -*-
"""
Created on Thu Jun 10 23:54:07 2021
@author: pgeir
"""

# TO BE IMPROVED --- Create API/Data Pipeline (Flask)

##########################################################################
# Import Working Directories & Libraries -------------------------------##
##########################################################################

import glob
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import pandas as pd
root = tk.Tk()
root.withdraw()

##########################################################################
# Functions ------------------------------------------------------------##
##########################################################################

def path_selection():
    folder = glob.glob(filedialog.askdirectory() + "/*.xlsx")
    while folder == []:
        print("ERROR: Folder chosen does not contain Excel Files.")
        glob.glob(filedialog.askdirectory() + "/*.xlsx")
    return folder

def excel_sheet_extraction(folder:str, sheets_list:list):
    sheets_list = []
    for file in folder:
        excel_wb = load_workbook(file, read_only=True, keep_links=False)
        sheets_list.extend(excel_wb.sheetnames)
    return list(set(sheets_list)) # Removes duplicates.

def keyword_selection(name:str):
    keyword = str(input(f"Type the keyword that identifies all the {name} Excel Sheets.\nKeyword: ")).lower().strip()
    return str(keyword)

# New list with the Excel Sheets that match a keyword only (Output => List).
def sheet_list_creation(keyword:str, filtered_sheet_list:list):
    filtered_sheet_list = [sheet_name for sheet_name in sheets_list if keyword in sheet_name.lower()]
    return list(filtered_sheet_list)

def dataframe_list_creation(filtered_sheet_list:list, df_list:list):
    df_list = []
    for file in folder:
        for sheet in filtered_sheet_list:
            try:
                data = pd.read_excel(file, sheet_name=sheet, header=2, usecols=("A:Q"))
                df_list.append(data)
            except ValueError:
                pass
    return list(df_list)

# Generates a concatenated Dataframe & respective Excel Files (Output => Excel File).
def concatenation(df_list:list, name:str):
    for df in df_list: # Give the same column names to all Dataframes.
        df.columns = df_list[0].columns
    df_list = pd.concat(df_list, axis=0, ignore_index=True, copy=False) # Concatenate all Dataframes.
    mask = df_list.iloc[:,1].isna() # Remove empty values.
    df_list = df_list.loc[~mask,:]
    save_file = f"{path_save}/ficheirolino_{name}.xlsx"
    df_list.to_excel(save_file, index=False)

def run_script(name:str, filtered_sheet_list:list, df_list:list):
    keyword = keyword_selection(name)
    filtered_sheet_list = sheet_list_creation(keyword, filtered_sheet_list)
    # While loop checks if the keyword string entered by the user is valid.
    while len(keyword) == 0 or filtered_sheet_list == []:
        print(f"ERROR: You must enter a keyword for the {name} data that matches the respective Sheet Names.")
        keyword = keyword_selection(name)
        filtered_sheet_list = sheet_list_creation(keyword, filtered_sheet_list)
    df_list = dataframe_list_creation(filtered_sheet_list, df_list)
    return concatenation(df_list, name)

#########################################################################
# Runnig Script -------------------------------------------------------##
#########################################################################

folder = path_selection() # Path of Input Excel files (user input).
path_save = filedialog.askdirectory() # Path of Output Excel files (user input).

# Create new Excel files
sheets_list = excel_sheet_extraction(folder, sheets_list="sheets_list")
run_script("Salesman", filtered_sheet_list="salesman_list", df_list="salesman_df")
run_script("Point_of_Sales", filtered_sheet_list="pos_list", df_list="pos_df")